# coding=utf-8
from openpyxl import Workbook
from openpyxl import load_workbook

if __name__ == '__main__':
    wb = load_workbook('./test.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')
    acount_list = []
    list = []
    list.append(sheet.cell(row=1,column=1).value)
    list.append(sheet.cell(row=1,column=2).value)
    for r in range(2, sheet.max_row + 1):
        acount_dict = {}
        acount_dict[list[0]] = sheet.cell(row=r,column=1).value
        acount_dict[list[1]] = sheet.cell(row=r,column=2).value
        acount_list.append(acount_dict)
    print(acount_list)
