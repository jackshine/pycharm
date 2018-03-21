# coding=utf-8
from openpyxl import Workbook
from openpyxl import load_workbook

if __name__ == '__main__':
    wb = load_workbook('./test.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')
    for r in range(1, sheet.max_row + 1):
        if(r == 1):
            continue
            # print('\n' + ''.join([str(sheet.cell(row=r, column=c).value).ljust(17) for c in range(1, sheet.max_column + 1)]))
        else:
            print(''.join([str(sheet.cell(row=r, column=c).value).ljust(20) for c in range(1, sheet.max_column + 1)]))





        # list = []
# print()
# # for i in range(sheet.):
# b2 = sheet.cell(row=2, column=1).value
# b3 = sheet.cell(row=3, column=1).value
# list.append(b2)
# list.append(b3)
# print(list)
