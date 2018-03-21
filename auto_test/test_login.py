from  selenium import webdriver
from openpyxl import load_workbook
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import time

def getFireFoxHandler():
    browser = webdriver.Firefox()
    return browser


def getIeHandler():
    browser = webdriver.Ie()
    return browser


def openUrl(b, url):
    b.get(url)


def readExcel():
    wb = load_workbook('./test.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')
    acount_list = []
    list = []
    list.append(sheet.cell(row=1, column=1).value)
    list.append(sheet.cell(row=1, column=2).value)
    for r in range(2, sheet.max_row + 1):
        acount_dict = {}
        acount_dict[list[0]] = sheet.cell(row=r, column=1).value
        acount_dict[list[1]] = sheet.cell(row=r, column=2).value
        acount_list.append(acount_dict)
    return acount_list


def waitEleById(b, id):
    locator = (By.ID, id)
    try:
        WebDriverWait(b, 20, 0.5).until(EC.visibility_of_element_located(locator))
    finally:
        print(locator)


def findEle(b, ele_id_list):
    # 找到元素
    waitEleById(b, ele_id_list['mobile_id'])
    name_Ele = b.find_element_by_id(ele_id_list['mobile_id'])
    pwd_Ele = b.find_element_by_id(ele_id_list['pwd_id'])
    login_Ele = b.find_element_by_id(ele_id_list['login_id'])
    Ele_list = []
    Ele_list.append(name_Ele)
    Ele_list.append(pwd_Ele)
    Ele_list.append(login_Ele)
    return Ele_list


def sendInfo(acount, Ele_list):
    list_key = ['mobile', 'password']
    for i in range(len(list_key)):
        Ele_list[i].send_keys("")
        Ele_list[i].send_keys(acount[list_key[i]])
    Ele_list[len(list_key)].click()
    return Ele_list


def checkResult(b):
    time.sleep(3)
    strs = ''
    print(b.current_url)
    if b.current_url == 'http://115.29.142.212:8010/Home/BookPage/index.html':
        strs = 'success login'
        b.quit()
    else:
        strs = b.find_element_by_id('login-tip').text
        b.quit()
    return strs  # def writeExcle(strs):
#

def controller(url, ele_id_dict):
    acount_list = readExcel()
    for acount in acount_list:
        print(acount)
        b = getFireFoxHandler()
        openUrl(b, url)
        Ele_list = findEle(b, ele_id_dict)
        sendInfo(acount, Ele_list)
        strs = checkResult(b)
        print(strs)
        # 写入excel
        # writeExcle(strs)


if __name__ == "__main__":
    url = "http://115.29.142.212:8010/login.html"
    ele_id_dict = {"mobile_id": "requestUsername", "pwd_id": "requestPassword", "login_id": "requestSubmit"}
    # login_test(acount_list, ele_id_dict)
    controller(url, ele_id_dict)
