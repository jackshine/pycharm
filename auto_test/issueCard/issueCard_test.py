# coding=utf-8
from selenium import webdriver
from openpyxl import load_workbook
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
import time
import os


def getFireFoxHandler():
    browser = webdriver.Firefox()
    return browser


def getIeHandler():
    browser = webdriver.Ie()
    return browser


def openUrl(b, url):
    b.get(url)


def findEle(b, ele_id_list):
    # 找到元素
    waitEleById(b, ele_id_list['mobile_id'])
    name_Ele = b.find_element_by_id(ele_id_list['mobile_id'])
    pwd_Ele = b.find_element_by_id(ele_id_list['pwd_id'])
    login_Ele = b.find_element_by_id(ele_id_list['login_id'])
    Ele_list = []
    Ele_list.append(name_Ele)
    Ele_list.append(pwd_Ele)
    Ele_list.append(login_Ele)

    return Ele_list


def sendInfo(acount_list, Ele_list):
    list_key = ['account_name', 'account_pwd']
    for i in range(len(list_key)):
        Ele_list[i].send_keys("")
        Ele_list[i].send_keys(acount_list[list_key[i]])
    Ele_list[len(list_key)].click()


def waitEle(b, xpath):
    locator = (By.XPATH, xpath)
    try:
        WebDriverWait(b, 10, 0.5).until(EC.visibility_of_element_located(locator))
    finally:
        print(locator)


def waitEleById(b, id):
    locator = (By.ID, id)
    try:
        WebDriverWait(b, 20, 0.5).until(EC.visibility_of_element_located(locator))
    finally:
        print(locator)


count = 1


def changeHotel(b):
    # 切换客栈
    # waitEle(b, '//*[@id="doc-header-brand"]/div')
    # change_Hotel_Ele = b.find_element_by_xpath('//*[@id="doc-header-brand"]')
    waitEle(b, '//*[@id="navbar-hotel-switch"]')
    change_Hotel_Ele = b.find_element_by_xpath('//*[@id="navbar-hotel-switch"]')

    # 点击客栈//*[@id="navbar-hotel-switch"]
    change_Hotel_Ele.click()

    # 若切换客栈该方法依然不行，则退出客栈

    # 若报异常，则又执行刷新页面，重新执行changeHeltel的方法
    global count
    print(count)
    if count == 2:
        loginOut(b)
        return False
        # 登录
    try:
        waitEle(b, '//*[@id="doc-header-brand-dropdown"]/div/li[4]/a')
        time.sleep(2)
        locationHotleEle = b.find_element_by_xpath('//*[@id="doc-header-brand-dropdown"]/div/li[4]/a')
        locationHotleEle.click()
        return True
    except TimeoutException as msg:
        print(msg)
        b.refresh()
        count += 1
        changeHotel(b)
    except NoSuchElementException as msg:
        print(msg)
        b.refresh()
        count += 1
        changeHotel(b)


# 办理入住
def issueCard(b):
    waitEle(b, "//*[@id='orderListBody']/tr[1]/td[4]/div")
    # 选择方格
    select_square_Ele = b.find_element_by_xpath("//*[@id='orderListBody']/tr[1]/td[4]/div")
    select_square_Ele.click()

    # 找“签发RF卡”元素
    waitEle(b, '//*[@id="orderInitContent"]/div[2]/div[1]/div[3]')
    issue_Ele = b.find_element_by_xpath('//*[@id="orderInitContent"]/div[2]/div[1]/div[3]')
    issue_Ele.click()

    time.sleep(2)
    # 获取签发的次数
    countEle_first = b.find_element_by_xpath('//*[@id="orderRfContainer"]/div/div[2]/div[2]/span').text
    print(countEle_first)

    # 签发RF卡
    waitEle(b, '//*[@id="orderRfContainer"]/div/div[2]/div[3]/div[1]')
    issue_button = b.find_element_by_xpath('//*[@id="orderRfContainer"]/div/div[2]/div[3]/div[1]')
    issue_button.click()

    time.sleep(5)
    # 获取点击签发后的次数
    countEle_second = b.find_element_by_xpath('//*[@id="orderRfContainer"]/div/div[2]/div[2]/span').text
    print(countEle_second)
    data = {}
    if int(countEle_first) + 1 == int(countEle_second):
        data['result'] = '签发成功'
    else:
        data['result'] = '签发失败'
    time.sleep(5)

    # 点击X按钮
    cancel_Ele = b.find_element_by_xpath('//*[@id="orderQRContent"]/div[1]/button')
    cancel_Ele.click()

    return data


def loginOut(b):
    time.sleep(3)
    url = "http://115.29.142.212:8010/logout.html"
    openUrl(b, url)


def writeExcel(data):
    wb = load_workbook('./test.xlsx')
    sheet = wb.get_sheet_by_name('Sheet2')
    temp = sheet.cell(row=2, column=2)
    temp.value = data['result']
    temp.font = Font(color=colors.BLACK)
    wb.save('./test.xlsx')




def controller(acount_list, ele_id_dict):
    # b = getFireFoxHandler()
    b = getIeHandler()
    url = 'http://115.29.142.212:8010/login.html'
    openUrl(b, url)
    # 登录
    Ele_list = findEle(b, ele_id_dict)
    sendInfo(acount_list, Ele_list)
    flag = changeHotel(b)
    if flag == False:
        controller(acount_list, ele_id_dict)
    else:
        data = issueCard(b)
        writeExcel(data)
        loginOut(b)


if __name__ == "__main__":
    account_name = '18926368199'
    account_pwd = "111111b"
    url = "http://115.29.142.212:8010/login.html"
    acount_list = {"account_name": account_name, "account_pwd": account_pwd}
    ele_id_dict = {"mobile_id": "requestUsername", "pwd_id": "requestPassword", "login_id": "requestSubmit"}
    # login_test(acount_list, ele_id_dict)
    controller(acount_list, ele_id_dict)
