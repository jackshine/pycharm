from  selenium import webdriver
from openpyxl import load_workbook
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import time
from openpyxl.styles import colors
from openpyxl.styles import Font, Color

def getFireFoxHandler():
    browser = webdriver.Firefox()
    return browser


def getIeHandler():
    browser = webdriver.Ie()
    return browser


def openUrl(b, url):
    b.get(url)
    time.sleep(3)


def readExcel():
    wb = load_workbook('./test.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')
    acount_list = []
    list = []
    list.append(sheet.cell(row=1, column=1).value)
    list.append(sheet.cell(row=1, column=2).value)
    for r in range(2, sheet.max_row + 1):
        acount_dict = {}
        acount_dict[list[0]] = sheet.cell(row=r,column=1).value
        if acount_dict[list[0]] == None:
            acount_dict[list[0]] = ''
        acount_dict[list[1]] = sheet.cell(row=r,column=2).value
        if acount_dict[list[1]] == None:
            acount_dict[list[1]] = ''
        acount_list.append(acount_dict)
    return acount_list

# 读取预期结果
def readExpectedResultsExcel():
    wb = load_workbook('./test.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')
    ExpectedResults = []
    for r in range(2, sheet.max_row + 1):
        ExpectedResults.append(sheet.cell(row=r, column=3).value)
    return ExpectedResults

def readEleByText():
    with open('./element_id.txt') as fb:
        dict = {}
        for line in fb:
            line_text = line.split('=')
            dict[line_text[0]] = line_text[1].replace('\n','')
    return dict

# 将结果写入excel中
def writeExcle(ActualResults):
    wb = load_workbook('./test.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')
    ExpectedResults = readExpectedResultsExcel()
    for i in range(ActualResults.__len__()):
        sheet.cell(row=i+2, column=4).value = ActualResults[i]
        temp = sheet.cell(row=i+2, column=5)
        if ActualResults[i] == ExpectedResults[i]:
             temp.value= 'pass'
             temp.font = Font(color=colors.BLACK)
        else:
             temp.value = 'fail'
             temp.font = Font(color=colors.RED)
    wb.save('./test.xlsx')

def waitEleById(b, id):
    locator = (By.ID, id)
    try:
        WebDriverWait(b, 20, 0.5).until(EC.visibility_of_element_located(locator))
    finally:
        print(locator)


def findEle(b, ele_id_list):
    # 找到元素
    waitEleById(b, ele_id_list['mobile_id'])
    name_Ele = b.find_element_by_id(ele_id_list['mobile_id'])
    pwd_Ele = b.find_element_by_id(ele_id_list['pwd_id'])
    login_Ele = b.find_element_by_id(ele_id_list['login_id'])
    Ele_list = []
    Ele_list.append(name_Ele)
    Ele_list.append(pwd_Ele)
    Ele_list.append(login_Ele)
    return Ele_list


def sendInfo(acount, Ele_list):
    list_key = ['mobile', 'password']
    for i in range(len(list_key)):
        Ele_list[i].send_keys("")
        Ele_list[i].send_keys(acount[list_key[i]])
    Ele_list[len(list_key)].click()
    return Ele_list

def loginOut(b):
    time.sleep(5)
    url ="http://115.29.142.212:8010/logout.html"
    openUrl(b,url)

def checkResult(b,url):
    time.sleep(3)
    strs = ''
    print(b.current_url)
    if b.current_url == 'http://115.29.142.212:8010/Home/BookPage/index.html':
        strs = '通过'
    else:
        strs = b.find_element_by_id('login-tip').text
    return strs  # def writeExcle(strs):
#


def controller(url, ele_id_dict):
    acountList = readExcel()
    ActualResults = []
    b = getIeHandler()
    openUrl(b, url)
    for acount in acountList:
        print(acount)
        Ele_list = findEle(b, ele_id_dict)
        sendInfo(acount, Ele_list)
        str = checkResult(b,url)
        ActualResults.append(str)
        if str == '通过':
            loginOut(b)
        else:
             b.refresh()
        # 写入excel
    writeExcle(ActualResults)


if __name__ == "__main__":
    url = "http://115.29.142.212:8010/login.html"
    ele_id_dict = readEleByText()
    controller(url, ele_id_dict)
