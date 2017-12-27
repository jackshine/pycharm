import random 
import string
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.styles import Border, Side, Font,colors 
import time
account = int(input("输入次数"))
wb = Workbook()
sheet = wb.create_sheet('Data', index=0) 
seed = "1234567890abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ"
sa = []
file_name = time.strftime("%Y%m%d%H%M%S",time.localtime())
count = 1
arr = []
for i in range(account):
	salt = ''
	sa = []
	for j in range(12):
		sa.append(random.choice(seed))
		salt = ''.join(sa)
	arr.append(salt)
	if(count%10==0):
		sheet.append(arr)
		arr = []
	count += 1
wb.save('./'+file_name+'.xlsx')
