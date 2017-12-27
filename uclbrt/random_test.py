import random 
import string
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
import time
account = int(input("输入次数"))
wb = Workbook()
sheet = wb.create_sheet('Data', index=0) 
seed = "1234567890abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ"
sa = []
file_name = time.strftime("%Y%m%d%H%M%S",time.localtime())
for i in range(account):
	salt = ''
	sa = []
	for j in range(12):
		sa.append(random.choice(seed))
		salt = ''.join(sa)
	sheet.append([salt])
wb.save('./'+file_name+'.xlsx')
