from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter

def write_excel():
    ex=load_workbook(filename=r'./case.xlsx')
    print('open excel success!')
    ws = ex.get_sheet_by_name("Sheet1")
    print('open sheet1 success!')
    ws.cell(row=4, column=2).value = 'hupi2222'
    print('write values success!')
    ex.save(filename='case.xlsx')
    print('save success!')


if __name__ == '__main__':
    write_excel()