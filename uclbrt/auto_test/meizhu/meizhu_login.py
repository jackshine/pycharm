# coding=utf-8
import unittest
from openpyxl import load_workbook
import requests
class Test(unittest.TestCase):
    def setUp(self):
        self.meizhu_client = 'http://115.29.142.212:8010'
        print('start')
    def test_case1(self):
        # 读取账号
        rb = load_workbook(filename='./meizhu_account.xlsx')
        table = rb.get_sheet_by_name("Sheet1")
        row_len = len(list(table.rows))
        for i in range(1,row_len):
            m_value = table.cell(row=i + 1, column=1).value
            p_value = table.cell(row=i + 1, column=2).value
            a_value = table.cell(row=i + 1, column=3).value
            if(m_value is  None):
                mobile = ''
            else:
                mobile = str(int(m_value))
            if(p_value is None):
                password = ''
            else:
                password = str(p_value)
            if(p_value is None):
                areaCode = ''
            else:
                areaCode = str(int(a_value))
            data = {
                'mobile':mobile,
                'password':password,
                'areaCode':areaCode
            }
            print(data)
                # 请求网络 # http://115.29.142.212:8010/Home/Public/login
            resp = requests.post(self.meizhu_client+'/Home/Public/login',data=data).json()
            # 根据网络返回的状态判断是否成功
            try:
                self.assertEqual(resp['status'], 200, msg='success')
                table.cell(row=i+1, column=4).value = resp['info']
            except:
                table.cell(row=i+1, column=4).value = resp['info']
            rb.save('./meizhu_account.xlsx')

    def tearDown(self):
        print("end")
if __name__ == "__main__":
    unittest.main









