import random 
import string
from openpyxl import load_workbook
account = int(input("输入次数"))
print(account)
rb = load_workbook(filename='./random.xlsx')
table = rb.get_sheet_by_name("Sheet1")
seed = "1234567890abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ"
#seed = "1234567890"
sa = []

for i in range(account):
	salt = ''
	sa = []
	for j in range(12):
		sa.append(random.choice(seed))
		salt = ''.join(sa)
	# table.cell(row=i+1, column=1).value = salt
	# rb.save('./random.xlsx')
	print(salt)

