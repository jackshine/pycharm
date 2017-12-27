from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter

dest_filename = 'empty_book.xlsx'

wb = Workbook()
ws1 = wb.active
ws1.title = "range names"
for row in range(1, 40):
   ws1.append(range(600))

ws3 = wb.create_sheet(title="Data")
for row in range(10, 20):
   for col in range(27, 54):
       _ = ws3.cell(column=col, row=row, value="%s" % get_column_letter(col))
print(ws3['AA10'].value)
wb.save(filename = dest_filename)

sheet_ranges = wb['range names']
print(sheet_ranges['D18'].value)

ws['A1'] = datetime.datetime(2010, 7, 21)
ws['A1'].number_format #输出'yyyy-mm-dd h:mm:ss'

rows = [
    ['Number', 'Batch 1', 'Batch 2'],
    [2, 40, 30],
    [3, 40, 25],
    [4, 50, 30],
    [5, 30, 10],
    [6, 25, 5],
    [7, 50, 10],
]

rows = [
    ['Date', 'Batch 1', 'Batch 2', 'Batch 3'],
    [date(2015,9, 1), 40, 30, 25],
    [date(2015,9, 2), 40, 25, 30],
    [date(2015,9, 3), 50, 30, 45],
    [date(2015,9, 4), 30, 25, 40],
    [date(2015,9, 5), 25, 35, 30],
    [date(2015,9, 6), 20, 40, 35],
]

for row in rows:
    ws.append(row)